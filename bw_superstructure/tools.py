import pandas as pd
import datetime
import pathlib as pt
from openpyxl import load_workbook
from typing import Optional

import brightway2 as bw

from bw_superstructure.bwutils.strategies import relink_exchanges_existing_db
from bw_superstructure.superstructure.utils import SUPERSTRUCTURE


def get_dbname_from_excel(path_to_excel):
    wb = load_workbook(path_to_excel)
    ws = wb.active
    for idx, icell in enumerate(ws["A"], 1):
        if icell.value.lower().strip() == "database":
            return ws.cell(row=idx, column=2).value
    raise IOError("no database name specified in column A&B of sheet {ws.name}")


def import_database(database_name: str, file_type: str, path_to_db):

    allowed_filetypes = ["ecospold", "excel", "bw2package"]
    assert (
        file_type in allowed_filetypes
    ), f'The file type "{file_type}" is not supported for importing a new database. Allowed file types are: "{allowed_filetypes}"'

    if file_type == "bw2package":
        print(f"Database {database_name} not yet in project, it will be imported.")
        db = bw.BW2Package.import_file(path_to_db)
        # this writes already the DB. DB is silently overwritten, also if it is already existing
        print(f"Database was written as: {db[0].name}")

        return

    if file_type == "ecospold":
        print(f'Database "{database_name}" not yet in project, it will be imported.')
        db = bw.SingleOutputEcospold2Importer(path_to_db, database_name, use_mp=False)
    elif file_type == "excel":
        print(
            f'Excel-based database "{database_name}" not  yet in project, it will be imported.'
        )
        db = bw.ExcelImporter(path_to_db)
        dbname_in_excel = get_dbname_from_excel(path_to_db)
        assert (
            dbname_in_excel == database_name
        ), f'The excel db name "{dbname_in_excel}" is inconsistent with the db name provided in the function: "{database_name}"'

    db.apply_strategies()
    db.statistics()

    try:
        db.write_database()
    except Exception:
        print(
            f'Error while writing database: "{database_name}". Check whether some exchanges cannot be linked correctly as provided in "{path_to_db}". See excel file of unlinked exchanges:'
        )
        db.write_excel(only_unlinked=True)
        print("Traceback:")  # provides exception from BW for db.write_database()


def check_and_import_database(
    db_name: str,
    fp_import_db: Optional[pt.Path] = None,
    filetype_import_db: Optional[str] = None,
):
    #################### BG-DB check / create if not exists
    db_exists = db_name in bw.databases

    if fp_import_db:  # fp is provided
        if db_exists:
            raise IOError(
                f'New background database "{db_name}" exists already. Therefore, we do not force a reimport of the database into the project. Set "fp_import_new_bg_db" to "None" if you want to use the already existing DB. Otherwise, delete the existing DB manually to be able to reimport it.'
            )

        #  DB does not yet exist
        import_database(
            db_name,
            file_type=filetype_import_db,
            path_to_db=fp_import_db,
        )

    elif not db_exists:  # no fp is provided, and DB does not yet exist
        raise IOError(
            f'The selected new background database "{db_name}" is not existing in your project. Therefore, we cannot relink to it. '
            f"Existing DBs are: \n {list(bw.databases)}. \n You can import the DB into the project by: \n     1. providing a filepath "
            f'to an excel file of the DB for "fp_import_new_bg_db" in "{relink_database_to_new_background.__name__}()" \n     '
            f'2. executing the function "import_database(database_name, file_type, path_to_db)" \n     3. importing it manually e.g. through the activity-browser'
        )


def relink_database_to_new_background(
    db_name: str,
    db_name_relinked: str,
    db_name_old_bg: str,
    db_name_new_bg: str,
    relink_output_db_without_copying_it: bool = True,
    fp_import_new_bg_db: Optional[pt.Path] = None,
    filetype_import_new_bg_db: Optional[str] = None,
):

    check_and_import_database(
        db_name=db_name_new_bg,
        fp_import_db=fp_import_new_bg_db,
        filetype_import_db=filetype_import_new_bg_db,
    )

    if relink_output_db_without_copying_it:
        # relink already existing db db_name_relinked directly
        assert (
            db_name_relinked in bw.databases
        ), f"{db_name_relinked} not in databases: {bw.databases}"
        print(
            f"relinking an already existing DB (withouth copying it): {db_name_relinked}"
        )

    else:
        # copy original DB db_name and save with new name=db_name_relinked
        assert (
            db_name_relinked != db_name
        ), f"\n 'db_name_relinked' is the same as 'output_dbname' which is set to: {db_name}. Moreover, 'relink_output_db_without_copying_it' is set to False. This is a problem, since frits.b will delete the existing db called db_name_relinked, but at the same time needs to copy it, which is technically impossible. You have the following options: \n    a. Choose a different name for 'db_name_relinked' \n    b. Set 'relink_output_db_without_copying_it' to 'True', such that the already existing output DB called '{db_name}' is directly relinked without creating a copy of it. "

        if db_name_relinked in bw.databases:
            del bw.databases[db_name_relinked]
            print(f"deleted database: {db_name_relinked}")

        bw.Database(db_name).copy(db_name_relinked)
        print(f"created new DB to be relinked: {db_name_relinked}")

    new_bg_db = bw.Database(db_name_new_bg)  # target background DB we want to relink to

    fg_db = bw.Database(
        db_name_relinked
    )  # foreground DB to relink from old background to new background DB

    relink_exchanges_existing_db(fg_db, db_name_old_bg, new_bg_db)


def remove_cols_from_sdf(
    sdf: pd.DataFrame, export_to_excel=False, excel_fp=None
) -> pd.DataFrame:
    """This functions removes some columns in the SDF (scenario difference file), which are unwanted for the input into the LCA-calc objects by the activity-browser.

    Args:
        sdf (pd.DataFrame): scenario difference file (in a pd.DataFrame)
        export_to_excel (bool, optional): _description_. Defaults to False.
        excel_fp (_type_, optional): _description_. Defaults to None.

    Returns:
        pd.DataFrame: _description_
    """

    sdf = sdf[
        sdf.columns.difference(SUPERSTRUCTURE)
    ]  # the columns specified in SUPERSTRUCTURE are removed

    if export_to_excel:
        sdf.to_excel(excel_fp)
        print("export processed SDF file to: {excel_fp}")

    return sdf


def time_stamp(fmt="%Y-%m-%d_%H-%M"):
    """creates a time stamp of now, e.g. "2023-02-15_13-37"

    Args:
        fmt (str, optional): format. Defaults to "%Y-%m-%d_%H-%M".

    Returns:
        str: time stamp, e.g. 2023-02-15_13-37
    """
    return datetime.datetime.now().strftime(fmt)
